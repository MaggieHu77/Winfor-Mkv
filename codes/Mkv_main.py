# -*- coding:utf-8 -*-
# ! python3

from configparser import ConfigParser
from Mkv_constant import *
from os import path, makedirs, mkdir
import sys
import time
from tqdm import tqdm
# import win32com.client as wc
from openpyxl import load_workbook, Workbook
from xlrd import open_workbook
from Mkv_data2 import create_stocks, calc_params, close2return
from Mkv_optimize2 import optimizer
from Mkv_pointwise_dates import BTdate
from Mkv_pointwise_codes import BTcodes
from Mkv_eval import MkvEval
from datetime import date, datetime
from collections import OrderedDict
from tqdm import tqdm
import numpy as np
import traceback
import pandas as pd
from WindPy import w
from Mkv_windpy_check import WindPyChecker
pd.options.mode.chained_assignment = None

"""
| 定义控制回测或实盘优化组合权重的入口类型Manager
| 通过读取配置文件configParam_mkv.conf获取用户设置
| 控制及部署任务：参数设置、数据读取、日期生成、循环数据、推送数据至优化器、收集并记录回测结果、写入文件
"""


class Manage:
    """
    管理及部署任务、调用各子模块、获取各子任务的结果；储存写出回测结果.
    """
    def __init__(self):
        """
        设定运行基本信息
        """
        self.code_dir = ""
        self.work_dir = ""
        self.target_index = ""
        self.global_spec = ""
        self.benchmark = ""
        self.calc_t = {}
        self.mode = 2
        self.frequency = "M"
        # only effective to hourly frequency,
        # instruct how many hour K to rebalance the portfolio
        self.rebalance_hour = None
        self.calendar = "SZSE"
        self.start_t = ""
        self.end_t = ""
        # self.first_t = ""
        self.num_d = T
        self.basic_indices = ""
        self.refresh_freq = ""
        self.up = DEFAULT_UP
        self.cash_r = CASH_RETURN
        self.cons_vol = DEFAULT_VOL / N ** 0.5
        self.calc_now = False
        self.short = None
        self.indices = ""
        self.ics_indices = {}
        self.codes = []
        self.stocks = []
        self.stocks_panel = {}
        self.t_seq = []
        self.t_eval_seq = []
        self.w = []
        self.w_panel = {}
        self.params = None
        self.mu_panel = {}
        self.cov_panel = {}
        self.bt_obj = None
        self.conf = ConfigParser()
        self.init_conf()
        self.read_target_index()
        self.set_input_mode()
        self.read_work_file()
        self.read_mode()
        self.read_benchmark()
        self.read_cons_vol()
        self.read_cons_up()
        self.read_cash_return()
        self.read_num_d()
        self.read_short()
        self.conf.write(open(CONF_NAME, "w+"))

    # ------------------------
    # 检验参数配置文件有效性
    # ------------------------
    def init_conf(self):
        """
        初始化参数配置文件.

        """
        f = open(CONF_NAME, "a+")
        f.close()
        self.conf.read(CONF_NAME, encoding="gbk")
        # self.conf.read_file(codecs.open(CONF_NAME, "r", encoding="gbk"))
        for sec in CONF_DICT.keys():
            if not self.conf.has_section(sec):
                self.conf.add_section(sec)
                for opt in CONF_DICT[sec]:
                    self.conf.set(sec, opt, "")
            else:
                for opt in CONF_DICT[sec]:
                    if not self.conf.has_option(sec, opt):
                        self.conf.set(sec, opt, "")
        print(f"-*-欢迎使用Markovitz组合优化系统 {VERSION}-*-")
        print("正在读取参数文件...")

    # ----------------------------------
    # 从configuration文件中读取参数设定
    # ----------------------------------
    def read_target_index(self):
        """
        从configuration文件中读取跟踪的目标指数.
        """
        self.target_index = self.conf.get("dir", "target_index")
        if not self.target_index:
            self.read_global_spec()
        else:
            if (self.target_index[-2:].upper() not in ["SH", "SZ"]) and (self.frequency == "H"):
                print("FrequencyError: 小时级别运行仅支持A股指数")
                sys.exit(1)

    def read_global_spec(self):
        """
        | 从configuration文件中读取基本面过滤全局股票空间，通常是某个全市场指数
        | 多个全局控件用;分隔
        """
        self.global_spec = self.conf.get("filter", "global_spec")
        if not self.global_spec:
            self.read_code_file()
        else:
            if self.frequency == "H":
                print("FrequencyError: 小时级别运行不支持基本面筛选模式")
            else:
                self.global_spec = self.global_spec.strip(" ").split(";")

    def read_code_file(self):
        """
        从configuration中读取记录回测股票代码文件参数地址设定.

        .. note::
           该函数在选择其他设置股票池的方法时，不会被调用，因此设置该方式的文件参数可以缺省.
        """
        self.code_dir = self.conf.get("dir", "code_file")
        flag = True
        while flag:
            if not path.isfile(self.code_dir):
                self.code_dir = input(FILE_EXIST_MSG)
                sys.exit(1)
            else:
                flag = False
                self.conf.set("dir", "code_file", self.code_dir)
        return

    def set_input_mode(self):
        """
        设置确定组合初始股票池或范围的模式类型.
          * 追踪某个指数成分股-1.
          * 在某个市场或多个板块概念成分组合下采用基本面选股-2.
          * 选取用户指定的股票，通过设置文件地址读入-3.
        """
        if self.target_index:
            self.input_mode = 1
            self.calendar = w.wss(self.target_index, "exch_eng").Data[0][0]
        elif self.global_spec:
            self.input_mode = 2
            self.read_indices()
            self.read_refresh_freq()
            self.read_ics()
            self.read_ics_fv()
            self.read_ics_rank()
        elif self.code_dir:
            self.input_mode = 3
        else:
            self.input_mode = None
            raise Exception("InputModeError: Can't decide stock pools.")

    def read_work_file(self):
        """
        | 从configuration读入工作文件夹地址.
        | 该工作文件夹同于存放回测结果.
        """
        self.work_dir = self.conf.get("dir", "work_file")

        if not path.exists(self.work_dir):
            if "/Mkv_output" not in self.work_dir:
                self.work_dir += "/Mkv_output"
            makedirs(self.work_dir)
            print(f"该文件夹不存在，已为您新建{self.work_dir}")

        else:
            if "/Mkv_output" not in self.work_dir:
                self.work_dir += "/Mkv_output"
                if not path.exists(self.work_dir):
                    mkdir(self.work_dir)
                    self.conf.set("dir", "work_file", self.work_dir)

    def read_benchmark(self):
        """
        读取对比的市场指数.
        """
        self.benchmark = self.conf.get("performance", "benchmark")

    def read_mode(self):
        """
        读取并设置运行模式.
          * 回测模式-1.
          * 实盘模式-2.
        """
        self.mode = self.conf.get("constraints", "mode")
        assert self.mode.isdigit(), "BacktestModeError:模式值不能为空，期望1=回测模式，2=实盘模式"
        self.mode = int(self.mode)
        assert self.mode in [1, 2], "BacktestModeError:模式值只能是1或2"
        if self.mode == 1:
            self.read_time_interval()
        else:
            self.read_calc_time()

    def read_time_interval(self):
        """
        从configuration文件中读入回测区间（起始时间）.
        """
        self.start_t = self.conf.get("calculation", "start_time")
        if not self.start_t:
            self.start_t = DEFAULT_START_T
            print(f"Warning:未输入回测起始时间，采用默认时间{DEFAULT_START_T}")
            self.conf.set("calculation", "start_time", self.start_t)

        self.end_t = self.conf.get("calculation", "end_time")
        if not self.end_t:
            self.end_t = DEFAULT_END_T
            print(f"Warning:未输入回测截止时间，采用默认时间{DEFAULT_END_T}")
            self.conf.set("calculation", "end_time", self.end_t)

        self.read_frequency()

    def read_frequency(self):
        """读取回测频率."""

        self.frequency = self.conf.get("calculation", "frequency")
        if self.mode != 2:
            assert isinstance(self.frequency, str) and self.frequency.upper() in ["M", "W",
                                                                                  "D", "H"], \
                "输入回测频率'%s'不支持" % self.frequency
            if self.frequency == "H":
                self.read_rebalance_hour()

    def read_rebalance_hour(self):
        """
        | read and check 'rebalance_hour' from configParam_mkv.conf.
        | use to confirm how many hourly K to recalculate the optimized weights and rebalance the portfolio.
        | require: freqency == "H" and in back-test mode.
        | should be a positive integer.
        """
        self.rebalance_hour = self.conf.get("calculation", "rebalance_hour")
        assert self.rebalance_hour.isdigit() and int(self.rebalance_hour) > 0, \
            "ParameterError: 'rebalance_hour' supposed to be a positive integer"
        self.rebalance_hour = int(self.rebalance_hour)

    def read_calc_time(self):
        """
        | 设置当日盘中价格截止时间.
        | 在实盘模式下有意义.
        """
        self.calc_time = self.conf.get("calculation", "calc_time")
        if not self.calc_time:
                self.calc_now = True
        else:
            from re import split
            t = [i.strip() for i in split("[:\- ]", self.calc_time)]
            # 补齐数组长度，防止用户漏写
            t = t + ["00"] * (6 - len(t))
            self.calc_t = {"y": int(t[0]), "m": int(t[1]), "d": int(t[2]),
                           "H": int(t[3]), "M": int(t[4]), "S": int(t[5])}

    def read_cons_vol(self):
        """
        读取优化的年化波动率约束.
        """
        cons_vol = self.conf.get("constraints", "vol").strip()
        if not cons_vol:
            self.cons_vol = DEFAULT_VOL / N ** 0.5
        else:
            self.cons_vol = float(cons_vol) / N ** 0.5
            if self.frequency == "H":
                self.cons_vol /= len(HOUR_BAR) ** 0.5

    def read_cons_up(self):
        """
        读取最大仓位约束.
        """
        self.up = float(self.conf.get("constraints", "max_weight").strip(""))

    def read_cash_return(self):
        """
        读取年化现金收益率.
        """
        self.cash_r = float(self.conf.get("constraints", "cash_return").strip("")) / 360

    def read_num_d(self):
        """
        读取回测日期长度.
        """
        self.num_d = int(self.conf.get("calculation", "num_d"))

    def read_short(self):
        """
        读取回测读取做空限制.
          * 不做空-0.
          * 允许做空-1.
        """
        self.short = int(self.conf.get("constraints", "short"))

    def read_indices(self):
        """
        读取分行业基本面指标筛选条件字符串.
        """
        self.indices = self.conf.get("filter", "basic_indices")
        if self.indices == "industry_gics":
            ics = list(ICS2_DICT.keys())
            for ii in ics:
                cond_ = self.conf.get("filter", ii)
                if not cond_:
                    assert self.indices, f"ParameterError: {ii} " \
                                         f"基本面变量缺失，无法basic_indices替代，因为basic_indices缺失"
                    cond_ = self.indices
                self.ics_indices.update({ICS2_DICT[ii]: cond_})
        else:
            print("IcsWarning:行业分类标准不是默认标准（wind二级行业），采用此标准将忽略子行业内的筛选标准，采用全局标准。")
            self.ics_indices = self.indices

    def read_refresh_freq(self):
        """
        读取更新筛选股票池更新频率.
        """
        self.refresh_freq = self.conf.get("filter", "refresh_freq").strip("\n\r").upper()
        if self.refresh_freq:
            assert int(self.refresh_freq[0]) in [1, 2, 3, 4, 5, 6] and self.refresh_freq[1] == \
                   "M", \
                "Error: parameter refresh_freq is not supported, should be in 1M~6M"

    def read_ics(self):
        """
        读取二级行业分类标准.
        默认采用wind行业二级分类，但是允许使用其他分类标准。
        一旦采用其他分类标准，则暂时子行业单独筛选标准不能使用。仅使用全局标准。

        """
        self.ics = self.conf.get("filter", "ics").strip("\n\r").lower()
        if not self.ics:
            self.ics = DEFAULT_ICS
            print("Warning: 参数ics（行业分类标准）未正确输入，采用默认输入--industry_gics（WIND行业分类）")

    def read_ics_fv(self):
        """
        读取子行业内排名变量，暂时只支持单个变量排序.
        """
        self.ics_fv = self.conf.get("filter", "ics_fv").strip("\n\r").lower()
        if not self.ics_fv:
            self.ics_fv = DEFAULT_ICS_FV
        else:
            ics_fv = FIELDS_ALIAS_DICT.get(self.ics_fv, 0)
            if not ics_fv:
                print(f"Warning：ics_fv（二级行业内排序变量）{ics_fv}不在程序默认字典内，推断是用户自选输入wind字段，请谨慎确认输入合法性")
            else:
                self.ics_fv = ics_fv

    def read_ics_rank(self):
        """
        | 读取子行业内排名变量，暂时只支持单个变量排序.
        | 读入特定排序的股票.
          * 如果输入单个整数，取前该整数个股票.
          * 如果输入m:n，取第m到第n只股票.
          * 如果输入[a,b,c,d]，取第a，第b，第c，第d只股票.
        """
        self.ics_rank = self.conf.get("filter", "ics_rank").strip("\n\r").lower()
        if not self.ics_rank:
            self.ics_rank = ICS_RANK
            print(f"Warning: 参数ics_rank（单个行业入选股票数量）未正确输入，采用默认输入rank={ICS_RANK}")
        elif self.ics_rank.isdigit():
            self.ics_rank = int(self.ics_rank)
        elif ":" in self.ics_rank:
            h, t = list(map(int, self.ics_rank.strip(" ").split(":")))
            h = h - 1
            self.ics_rank = (h, t)
        elif "[" in self.ics_rank:
            rank_ = list(map(int, self.ics_rank.strip("[]").split(",")))
            self.ics_rank = list(map(lambda x: x-1, rank_))
        else:
            raise Exception("ParameterError: 参数ics_rank必须为正整数,索引切片m:n或者list [m,n,p,q]形式")

    def read_codes(self):
        """
        从指定文件中读取股票代码.
        """
        print("*************************************************************")
        self.codes = read_codes(self.code_dir)
        print(f"%读取{len(self.codes)}只股票")

    def set_params1(self):
        """
        | 回测模式的权重优化及回测结果计算、写出.
        | 根据不同设置初始组合控件的方法提取有效参数.
        | 创建Mkv_pointwise_codes.BTcodes类对象，用于生成回测期间需要进入组合优化的有效股票.
        | 遍历各个调仓日期，计算待配权股票在统计期内的均值和协方差矩阵.
        | 推送统计量和参数进入mosek进行Markowitz组合寻优.
        | 模拟回测期间的持仓并记录和写出回测结果.
        """
        if self.input_mode == 3:
            input_params = {"code_file": self.code_dir,
                            "calendar": self.calendar}
        elif self.input_mode == 1:
            input_params = {"target_index": self.target_index,
                            "calendar": self.calendar}
        else:
            input_params = {"global_spec": self.global_spec,
                            "basic_indices": self.indices,
                            "ics": self.ics,
                            "ics_fv": self.ics_fv,
                            "ics_rank": self.ics_rank,
                            "refresh_freq": self.refresh_freq,
                            "ics_indices": self.ics_indices,
                            "calendar": self.calendar}
        codes_object = BTcodes(input_mode=self.input_mode)
        codes_object.set_codes_env(params=input_params)
        for tt, month_id in self.t_seq:
            print(f"-正在回测时点{tt}")
            codes = codes_object.get_current_codes(date=tt, month_id=month_id)
            stocks, valid_codes = create_stocks(codes, self.num_d, self.frequency, self.calendar,
                                               tt, False)
            self.stocks_panel.update({tt: stocks})
            params = calc_params(stocks, self.short)
            params.update({"vol": self.cons_vol, "up": self.up, "cash_r": self.cash_r})
            self.mu_panel.update({tt: params["mu"]})
            self.cov_panel.update({tt: params["cov"]})
            w = optimizer(params)
            self.w_panel.update({tt: w})
            print(f"--Backtest[{tt}] output:{np.matmul(np.matmul(np.array(w), params['cov']), np.array(w).T)}")
            print(f"--constraint:{self.cons_vol**2}")
        self.t_seq = [tt[0] for tt in self.t_seq]
        self.t_eval_seq = self.bt_obj.get_eval_dates(ref_seq=self.t_seq)
        self.write_output1d(write_basic_K_rt=True)

    def set_params2(self, q=True):
        """
        实盘问题的计算优化参数，优化当前日期可得数据，下期持仓组合.

        .. deprecated:: v0.6.3
           该函数需要进行增强，当前函数版本功能不全.
        """
        if self.input_mode == 3:
            input_params = {"code_file": self.code_dir}
        elif self.input_mode == 1:
            input_params = {"target_index": self.target_index}
        else:
            input_params = {"global_spec": self.global_spec,
                            "basic_indices": self.indices,
                            "ics": self.ics,
                            "ics_fv": self.ics_fv,
                            "ics_rank": self.ics_rank,
                            "refresh_freq": self.refresh_freq,
                            "ics_indices": self.ics_indices}
        codes_object = BTcodes(input_mode=self.input_mode)
        codes_object.set_codes_env(params=input_params)
        self.codes = codes_object.get_current_codes(date=datetime.now().strftime("%Y-%m-%d"),
                                               month_id=0)
        if not q:
            self.stocks, self.codes = create_stocks(self.codes,
                                        self.num_d,
                                        self.frequency,
                                        self.calendar,
                                        datetime(*self.calc_t.values()).strftime("%Y-%m-%d"),
                                        q=q)
        else:
            self.stocks, self.codes = create_stocks(self.codes, self.num_d, self.frequency,
                                         calendar=self.calendar)
        self.params = calc_params(self.stocks, self.short)
        self.params.update({"vol": self.cons_vol, "up": self.up, "cash_r": self.cash_r})
        self.w = optimizer(self.params)
        print(
            f"output:"
            f"{np.matmul(np.matmul(np.array(self.w), self.params['cov']), np.array(self.w).T)}\n"
            f"constraint:{self.cons_vol**2}")
        self.write_output2()

    def _get_backtest_t_seq(self):
        """
        | 获取优化调仓时点序列.
        | 创建Mkv_pointwise_dates.BTdate类对象，用于生成准确调仓日期序列.
        """
        self.bt_obj = BTdate(start=self.start_t,
                             end=self.end_t,
                             freq=self.frequency,
                             rebalance_h=self.rebalance_hour,
                             calendar=self.calendar)
        self.t_seq = self.bt_obj.get_backtest_dates()
        # self.first_t = self.bt_obj.get_backtest_first_date(num_d=self.num_d)

    def run_optimizer(self):
        """
        | 根据运行模式的不同，选择回测或者是实盘函数入口.
        """
        if self.mode == 1:
            self._get_backtest_t_seq()
            print(f"-*-进入回测模式，检测到{len(self.t_seq)}个回测时点-*-")
            self.set_params1()
        else:
            stamp = datetime.now()
            calc_t = datetime(*self.calc_t.values())
            # 给定日期非当天，用于回测而非实盘
            if calc_t.date() < stamp.date():
                self.set_params2(q=False)
            # 给定日期在当天或者在未来，都当做在当天进行
            elif calc_t.date() > stamp.date():
                self.set_params2()
            else:
                if calc_t <= stamp:
                    self.set_params2()
                else:
                    delta_t = (calc_t - stamp).seconds
                    print("%定时等待运行...")
                    from time import sleep
                    sleep(delta_t)
                    self.set_params2()

    def write_output1(self, write_basic_K_rt=False):
        """
        | 将结果写入excel文件.

        .. deprecated:: v.0.6.3
           该函数未来将被移除.
           Use :func:`write_output1d` instead.

        :param write_basic_K_rt: 是否单独记录回测期间逐K线的组合收益率.

        """
        print("%正在写入结果到文件...")
        wb = Workbook()
        s1 = wb.active
        s1.title = "weights"
        s2 = wb.create_sheet()
        token = {"M": "monthly", "W": "weekly", "D": "daily", "H": "hourly"}[self.frequency]
        s2.title = "portfolio %s returns" % token
        s3 = wb.create_sheet()
        s3.title = f"{self.num_d}-day mean returns"
        s4 = None
        if write_basic_K_rt and (token not in ["daily", "hourly"]):
            s4 = wb.create_sheet()
            s4.title = f"portfolio daily returns"
        elif write_basic_K_rt and token == "hourly":
            print("Warning: 组合回测外期间日频收益率追踪暂不支持小时级别调仓")
        self.codes = [s.code for s in self.stocks_panel[self.t_seq[0]]]
        # 写第一张表回测各周期weights
        print(f"%写入sheet1...{s1.title}")
        col11 = ["code"] + self.codes + ["cash"]
        names = [s.name for s in self.stocks_panel[self.t_seq[0]]]
        col12 = ["name"] + names + ["现金"]
        content1 = [col11, col12]
        for tt in self.t_seq:
            content1.append([tt] + self.w_panel[tt])
        # 采取长表形式，时间为行索引
        # content1 = list(zip(*content1))
        for row in content1:
            s1.append(row)
        # 写第二张表回测权重持有至下一期的组合期间收益率
        print(f"%写入sheet2...{s2.title}")
        eval_obj = MkvEval(eval_seq=self.t_eval_seq,
                           freq=self.frequency,
                           stocks=self.codes,
                           cash_r=self.cash_r,
                           calendar=self.calendar)
        returns = eval_obj.eval_returns()
        from numpy import dot
        col22 = ["portfolio return(%)"] + \
                [dot(self.w_panel[self.t_seq[t]], returns[t]) for t in
                                         range(len(self.t_seq))]
        col21 = ["evaluation_date"] + self.t_eval_seq
        content2 = list(zip(col21, col22))
        for rr in content2:
            s2.append(rr)

        # 写第三张表
        content3 = [col11, col12]
        print(f"%写入sheet3...{s3.title}")
        for tt in self.t_seq:
            content3.append([tt] + self.mu_panel[tt] + [self.cash_r])
        # content3 = list(zip(*content3))
        for row in content3:
            s3.append(row)

        if s4 is not None:
            print(f"%写入sheet4...{s4.title}")
            # write portfolio holding daily return
            # call a MkvEval object to calculate daily returns
            eval_start_ = w.tdaysoffset(1, self.t_seq[0], f'''TradingCalendar={
                                                          self.calendar}''').Data[0][0].strftime("%Y-%m-%d")
            eval_end_ = self.t_eval_seq[-1]
            eval_obj2 = MkvEval(eval_seq=[eval_start_, eval_end_],
                                cash_r=self.cash_r,
                                freq="D",
                                stocks=self.codes,
                                calendar=self.calendar)
            returns2, eval_t_ = eval_obj2.eval_returns(get_t_seq=True)
            row_h = ["evaluation_date", "portfolio daily return(%)"]
            s4.append(row_h)
            pre_w = None
            pre_r = None
            tqdm_obj = tqdm(range(len(eval_t_)))
            for tt in tqdm_obj:
                # 该交易日初的权重
                cur_w = lookup_w(w_panel=self.w_panel, t=eval_t_[tt], pre_w=pre_w, pre_r=pre_r,
                                 calendar=self.calendar)
                cur_r = np.dot(cur_w, returns2[tt])
                s4.append([eval_t_[tt], cur_r])
                pre_w = cur_w
                pre_r = returns2[tt]

        # 写回测期间协方差的表
        # for t in range(len(self.t_seq)):
        #     exec(f"s{4+t}=wb.create_sheet()")
        #     exec(f"s{4+t}.title='Cov_{self.t_seq[t]}'")
        #     for row in self.cov_panel[self.t_seq[t]]:
        #         exec(f"s{4+t}.append({list(row)})")

        output_dir = f'''{self.work_dir}/mkv_{path.splitext(path.basename(self.code_dir))[0]}
_{['longOnly', 'short'][self.short]}_{''.join(self.start_t.split('-'))}-
{''.join(self.end_t.split('-'))}_freq{self.frequency}t.xlsx'''.replace("\n", "")
        wb.save(output_dir)
        print(f"%结果输出到：{output_dir}")

    def write_output1d(self, write_basic_K_rt=False):
        """
        将回测结果写入excel文件.
        """
        print("%正在写入结果到文件...")
        wb = Workbook()
        s1 = wb.active
        s1.title = "weights"
        s2 = wb.create_sheet()
        token = {"M": "monthly", "W": "weekly", "D": "daily", "H": "hourly"}[self.frequency]
        if token != "hourly":
            s2.title = "portfolio %s returns" % token
        else:
            s2.title = "portfolio %s%s returns" %(self.rebalance_hour, token)
        s3 = wb.create_sheet()
        s3.title = f"{self.num_d}-day mean returns"
        s4 = None
        if write_basic_K_rt and (token in ["monthly", "weekly"]):
            s4 = wb.create_sheet()
            s4.title = "portfolio daily returns"
        elif write_basic_K_rt and token == "hourly":
            s4 = wb.create_sheet()
            s4.title = "portfolio hourly returns"
        # 写第一张表回测各周期weights
        print(f"%写入sheet1...{s1.title}")
        content1, weight_df = dict2list_v2(stock_pannel=self.stocks_panel, var_pannel=self.w_panel,
                             date_first=True)
        for row in content1:
            s1.append(row)
        # 写第二张表回测权重持有至下一期的组合期间收益率
        print(f"%写入sheet2...{s2.title}")
        eval_obj = MkvEval(eval_seq=self.t_eval_seq,
                           freq=self.frequency,
                           stocks=[],
                           cash_r=self.cash_r,
                           calendar=self.calendar)
        from numpy import dot
        col22 = ["portfolio return(%)"]
        # m_df = pd.DataFrame()
        for t in range(len(self.t_eval_seq)):
            codes = [s.code for s in self.stocks_panel[self.t_seq[t]]]
            returns = eval_obj.eval_returns12(codes=codes,
                                              eval_date=self.t_eval_seq[t],
                                              eval_interval_start=self.t_seq[t],
                                              include_head=False)
            returns_seq = pd.Series(returns, index=codes+["cash"], name=t)
            weight_df = pd.concat([weight_df, returns_seq], axis=1)
            # m_df = pd.concat([m_df, pd.Series(returns, index=codes+["cash"], name=self.t_eval_seq[
            #     t])], axis=1)
            port_t = weight_df[self.t_seq[t]].fillna(0).dot(weight_df[t].fillna(0))
            # col22.append(dot(self.w_panel[self.t_seq[t]], returns))
            col22.append(port_t)
        # m_df.T.to_csv("r_monthly.csv")
        col21 = ["evaluation_date"] + self.t_eval_seq
        if self.benchmark:
            if self.frequency in ["M", "W", "D"]:
                col23 = [f"{self.benchmark} returns(%)"] + \
                        get_benchmark_returns(self.benchmark,
                                              time_seq=self.t_eval_seq,
                                              freq=self.frequency,
                                              percentage=True,
                                              tolist=True,
                                              calendar=self.calendar)
            else:
                col23 = [f"{self.benchmark} returns(%)"] + \
                        get_benchmark_returns(self.benchmark,
                                              time_seq=self.t_eval_seq,
                                              freq=self.frequency,
                                              percentage=True,
                                              tolist=True,
                                              rebalance_hour=self.rebalance_hour,
                                              pre_backtest_t=self.t_seq[0],
                                              calendar=self.calendar)
            content2 = list(zip(col21, col22, col23))
        else:
            content2 = list(zip(col21, col22))
        for rr in content2:
            s2.append(rr)
        # s2.append(["evaluation_date"] + self.t_eval_seq)
        # s2.append(col22)

        # 写第三张表
        print(f"%写入sheet3...{s3.title}", sep="\n")
        content3 = dict2list(stock_pannel=self.stocks_panel,
                             var_pannel=self.mu_panel,
                             suffix=[self.cash_r],
                             date_first=True)
        for row in content3:
            s3.append(row)

        # 写第四张表
        if s4 is not None:
            print(f"%写入sheet4...{s4.title}")
            count_start = time.time()
            eval_obj2 = MkvEval(eval_seq=self.t_eval_seq,
                                cash_r=self.cash_r,
                                freq=["D", "H"][self.frequency == "H"],
                                stocks=[],
                                calendar=self.calendar)

            if self.frequency == "H":
                print("%might take a little bit of time...")
                tqdm_obj = tqdm(range(len(self.t_eval_seq)))

                init_port = 1.0
                port_periods = []
                for t in tqdm_obj:
                    tqdm_obj.set_description(f"eval hourly returns: {self.t_eval_seq[t]}", False)
                    codes = [s.code for s in self.stocks_panel[self.t_seq[t]]]
                    returns2, eval_t_ = eval_obj2.eval_returns_hrange(
                        codes=codes,
                        eval_begin=self.t_seq[t],
                        eval_end=self.t_eval_seq[t],
                        skip=self.rebalance_hour,
                        get_t_seq=True)
                    port_per, init_port = get_portfolio_basic_K_valuess(rts_arr=returns2,
                                                                        time_seq=eval_t_,
                                                                        weights=self.w_panel[
                                                                            self.t_seq[t]],
                                                                        init_port_value=init_port,
                                                                        percentage=True)
                    port_periods.append(port_per)
                port_values = pd.concat(port_periods, axis=0).pct_change()
                port_values.iloc[0] = port_periods[0].iloc[0] - 1
                port_values = port_values * 100
                if self.benchmark:
                    benchmark_rt = get_benchmark_returns(self.benchmark,
                                                         time_seq=port_values.index.tolist(),
                                                         freq="H",
                                                         calendar=self.calendar,
                                                         percentage=True,
                                                         tolist=True,
                                                         rebalance_hour=1)
                    content4 = pd.DataFrame(port_values)
                    content4["benchmark"] = benchmark_rt
                    content4.reset_index(inplace=True)
                    content4 = content4.values.tolist()
                    content4.insert(0, ["evaluation_date", s4.title+"(%)", f"{self.benchmark} (%)"])
                else:
                    content4 = pd.DataFrame(port_values)
                    content4.reset_index(inplace=True)
                    content4 = content4.values.tolist()
                    content4.insert(0,
                                    ["evaluation_date", s4.title + "(%)", f"{self.benchmark} (%)"])

            else:
                tqdm_obj = tqdm(range(len(self.t_eval_seq)))
                init_port = 1.0
                port_periods = []
                for t in tqdm_obj:
                    tqdm_obj.set_description(f"eval daily returns: {self.t_eval_seq[t]}", False)
                    codes = [s.code for s in self.stocks_panel[self.t_seq[t]]]
                    returns2, eval_t_ = eval_obj2.eval_returns_drange(
                        codes=codes,
                        eval_begin=self.t_seq[t],
                        eval_end=self.t_eval_seq[t],
                        get_t_seq=True)
                    returns2.pop(0)
                    eval_t_.pop(0)
                    port_per, init_port = get_portfolio_basic_K_valuess(rts_arr=returns2,
                                                                        time_seq=eval_t_,
                                                                        weights=self.w_panel[
                                                                            self.t_seq[t]],
                                                                        init_port_value=init_port,
                                                                        percentage=True)
                    port_periods.append(port_per)
                port_values = pd.concat(port_periods, axis=0).pct_change()
                port_values.iloc[0] = port_periods[0].iloc[0] - 1
                port_values = port_values * 100
                if self.benchmark:
                    benchmark_rt = get_benchmark_returns(self.benchmark,
                                                         time_seq=port_values.index.tolist(),
                                                         freq="D",
                                                         calendar=self.calendar,
                                                         percentage=True,
                                                         tolist=True)
                    content4 = pd.DataFrame(port_values)
                    content4["benchmark"] = benchmark_rt
                    content4.reset_index(inplace=True)
                    content4 = content4.values.tolist()
                    content4.insert(0,
                                    ["evaluation_date", s4.title + "(%)", f"{self.benchmark} (%)"])
                else:
                    content4 = pd.DataFrame(port_values)
                    content4.reset_index(inplace=True)
                    content4 = content4.values.tolist()
                    content4.insert(0,
                                    ["evaluation_date", s4.title + "(%)", f"{self.benchmark} (%)"])
            for rr in content4:
                s4.append(rr)
            count_end = time.time()
            print(f"sheet4 cost {count_end-count_start} secs")
        # 写回测期间协方差的表
        # for t in range(len(self.t_seq)):
        #     exec(f"s{4 + t}=wb.create_sheet()")
        #     exec(f"s{4 + t}.title='Cov_{self.t_seq[t]}'")
        #     for row in self.cov_panel[self.t_seq[t]]:
        #         exec(f"s{4 + t}.append({list(row)})")
        rebalance_freq = self.frequency if self.frequency!="H" else str(self.rebalance_hour)+"H"
        output_dir = f'''{self.work_dir}/mkv_{[self.target_index, self.global_spec, 
        path.splitext(path.basename(self.code_dir))[0]][self.input_mode-1]}_{['longOnly', 'short'][self.short]}_
{''.join(self.start_t.split('-'))}-{''.join(self.end_t.split('-'))}_freq{
rebalance_freq}.xlsx'''.replace("\n", "")
        wb.save(output_dir)
        print(f"%结果输出到：{output_dir}")

    def write_output2(self):
        """
        将实盘模式的结果写入excel文件.
        """
        print("%正在写入结果到文件...")
        wb = Workbook()
        s1 = wb.active
        s1.title = "weights"
        s2 = wb.create_sheet()
        s2.title = f"{self.num_d}-days mean returns"
        # s3 = wb.create_sheet()
        # s3.title = "covariance"
        # 写第一张表-weight
        s1.cell(row=1, column=1, value="code")
        s1.cell(row=1, column=2, value="name")
        s1.cell(row=1, column=3, value="weight")
        for cc in range(len(self.stocks)):
            s1.cell(row=cc + 2, column=1, value=self.stocks[cc].code)
            s1.cell(row=cc + 2, column=2, value=self.stocks[cc].name)
            s1.cell(row=cc + 2, column=3, value=self.w[cc])
        s1.cell(row=len(self.w) + 1, column=1, value="cash")
        s1.cell(row=len(self.w) + 1, column=2, value="现金")
        s1.cell(row=len(self.w) + 1, column=3, value=self.w[-1])
        # 写第二张表-return
        s2.cell(row=1, column=1, value="code")
        s2.cell(row=1, column=2, value="name")
        s2.cell(row=1, column=3, value="mu")
        for cc in range(len(self.stocks)):
            s2.cell(row=cc + 2, column=1, value=self.stocks[cc].code)
            s2.cell(row=cc + 2, column=2, value=self.stocks[cc].name)
            s2.cell(row=cc + 2, column=3, value=self.params["mu"][cc])
            for t in range(len(self.stocks[0].t)):
                s2.cell(row=1, column=t + 4, value=self.stocks[0].t[t])
                s2.cell(row=cc + 2, column=t + 4, value=self.stocks[cc].r[t])
        s2.cell(row=len(self.w) + 1, column=1, value="cash")
        s2.cell(row=len(self.w) + 1, column=2, value="现金")
        s2.cell(row=len(self.w) + 1, column=3, value=0.)
        for t in range(len(self.stocks[0].t)):
            s2.cell(row=len(self.w) + 1, column=t + 4, value=self.cash_r)
        # 写第三张表-covariance
        # s3.cell(row=1, column=1, value="code")
        # s3.cell(row=1, column=2, value="name")
        # for cc in range(len(self.stocks)):
        #     s3.cell(row=cc + 2, column=1, value=self.stocks[cc].code)
        #     s3.cell(row=cc + 2, column=2, value=self.stocks[cc].name)
        # for i in range(len(self.params['qval'])):
        #     s3.cell(row=self.params['qsubi'][i] + 2,
        #             column=self.params['qsubj'][i] +3,
        #             value=self.params['qval'][i])
        output_dir = f"{self.work_dir}/mkv_{[self.target_index, self.global_spec][self.input_mode-1]}"+ \
            f"_{['longOnly', 'short'][self.short]}_{datetime(*self.calc_t.values()).strftime('%y%m%d')}.xlsx"
        wb.save(output_dir)
        print(f"%结果输出到：{output_dir}")


def read_codes(f):
    """
    从文件地址读入股票代码.

    :param f: 文件绝对路径.

    :return: list，组合成分股列表.
    """
    b_name = path.basename(f)
    b_type = b_name.split(".")[-1]
    codes = []
    # if b_type == 'xls':
    #     # openpyxl不支持xls格式，先转化为xls
    #     e = wc.gencache.EnsureDispatch('Excel.Application')
    #     wb = e.Workbooks.Open(f)
    #     wb.SaveAs(f + 'x', FileFormat=51)
    #     wb.Close()
    #     e.Application.Quit()
    #     f = f + 'x'
    #     b_type = 'xlsx'
    if b_type == "txt":
        for code in open(f):
            code = code.strip("\n|;|,|/|")
            codes.append(code)
    elif b_type == "xlsx":
        wb = load_workbook(f)
        sheet = wb.active
        for code in list(sheet.columns)[0]:
            if code.value and "." in code.value:
                codes.append(code.value)
    elif b_type == 'xls':
        wb = open_workbook(f)
        sheet = wb.sheet_by_index(0)
        for code in sheet.col_values(0):
            if "." in code:
                codes.append(code)
    else:
        print(f"Error: 不支持的文件格式{b_type}")
    return codes


# helper func: merge weights in different sets of stock pools
def dict2list(stock_pannel, var_pannel, suffix=None, date_first=False):
    """
    这个函数扩展性不佳，为了分隔较复杂的数据形式重组功能.

    :param stock_pannel: 包含回测全部权股票基本信息的字典数据.
    :param var_pannel: 包含回测全部权重信息的字典数据.
    :param suffix: 需要附加在var_pannel元素后的list.

    :return: 通过pd.DataFrame merge各个不同日期回测结果后转换为可以直接写入workbook的list形式.
    """
    # 提取时间索引
    t_seq = list(stock_pannel.keys())
    t_start = t_seq[0]

    # 定义函数将每次回测结果转化为DataFrame
    def dict2dataframe(key):
        content0 = [[s.code, s.name] for s in stock_pannel[key]]
        content0.append(["cash", "现金"])
        content0 = dict(zip(["code", "name"],zip(*content0)))
        if suffix:
            if isinstance(var_pannel[key], list):
                var_pannel_ = var_pannel[key] + suffix
            else:
                var_pannel_ = list(var_pannel) + suffix
            content0.update({key: var_pannel_})
            df0 = pd.DataFrame(content0)
        else:
            content0.update({key: var_pannel[key]})
            df0 = pd.DataFrame(content0)
        return df0

    # 初始化总表
    df = dict2dataframe(t_start)
    # 通过循环，每次加入一次回测结果，保存之前结果merge后的数据框
    for tt in t_seq[1:]:
        df0 = dict2dataframe(tt)
        df = pd.merge(df, df0, on=["code", "name"], how="outer", copy=False)
    # 由于没有采用index，需要在输出之前调整行列顺序使其美观
    # 调整列顺序
    df = df[["code", "name"] + t_seq]
    # 调整行序，使得现金始终排在最后一行
    cash_id = df[df.code == "cash"].index.tolist()
    cash_row = df.iloc[cash_id]
    df.drop(index=cash_id, inplace=True)
    df = df.append(cash_row)
    df2list = df.values.tolist()
    df2list.insert(0, df.columns.tolist())
    if date_first:
        df2list = list(zip(*df2list))
    return df2list


def dict2list_v2(stock_pannel, var_pannel, suffix=None, date_first=False):
    out_df = pd.DataFrame()
    for tt, ss in stock_pannel.items():
        codes = [s.code for s in ss] + ["cash"]
        ww = pd.Series(var_pannel[tt], name=tt, index=codes)
        out_df = pd.concat([out_df, ww], axis=1)
    out_df.fillna(0, inplace=True)
    content = out_df.T.values.tolist()
    codes = ["code"] + out_df.index.tolist()
    dates = out_df.columns.tolist()
    for ii in range(len(content)):
        content[ii] = [dates[ii]] + content[ii]
    content.insert(0, codes)
    return content, out_df


def lookup_w(w_panel, t, calendar, pre_w=None, pre_r=None):
    """
    用于返回对应高频（这里是daily）日期在回测期间的权重（调频周期为低频周、月）.

    :param w_panel: dict，在调仓日（收盘后）的对应新权重.
    :param t: str, 需要确定组合日收益率的当日.
    :param pre_w: list, 前2交易日收盘后的权重，即是前1交易日开盘前权重.
    :param pre_r: list, 前1交易日当天收益率.

    :return: list,t当日开盘前权重.

    """
    assert isinstance(w_panel, dict), "TypeError: w_panel should be type dict"
    # 按时间顺序排序
    w_panel = OrderedDict(sorted(w_panel.items(), key=lambda k: k[0]))
    rebalance_seq = list(w_panel.keys())
    # 由上期莫转换到本期初
    try:
        tt = rebalance_seq[0]
        rebalance_seq_ = list(map(lambda tt: w.tdaysoffset(1, tt, f'''TradingCalendar={
                                                                  calendar}''').Data[0][0].strftime(
            "%Y-%m-%d"),
                                 rebalance_seq))
    except AttributeError as e:
        print(f"{tt} {e}")
    if t[:10] in rebalance_seq_:
        idx = rebalance_seq_.index(t[:10])
        cur_w = w_panel[rebalance_seq[idx]]

    else:
        # 价格变化导致每天的实际权重发生变化
        cur_w = list(np.array(pre_w) * (np.array(pre_r)/100 + 1) / np.dot(np.array(pre_r)/100 + 1,
                                                                 np.array(pre_w)))
    return cur_w


def get_benchmark_returns(benchmark, time_seq, freq, calendar, percentage, tolist,
                          rebalance_hour=1, pre_backtest_t=""):
    """
    获得给定频率和时间内的benchmark收益率数据.用close计算。

    :param benchmark: str, 通常是某市场指数
    :param time_seq: list, 回测期间待对比的时间点
    :param freq: str, 应当是"M","W","D","H"之一
    :param calendar: str, 交易所日历，应当使用组合的基准日历而不是benchmark的市场日历
    :param percentage: bool, 是否返回百分数形式的收益率
    :param tolist: bool, 是否以list形式返回，False则返回pd.Series
    :param rebalance_hour: int, 特别针对freq="H"的情况，指调仓的频率
    :param pre_backtest_t: str, 特别针对rebalance_hour >1的情况，表示上一个调仓时间

    :return: list, pd.Series，benchmark的收益率序列
    """

    if freq in ["M", "W", "D"]:
        start_ = w.tdaysoffset(-1, time_seq[0],
                               f"Period={freq};TradingCalendar={calendar}").Data[0][0].strftime(
            "%Y-%m-%d")
        end_ = time_seq[-1]
        res = w.wsd(benchmark, "close", start_, end_,
                    f"Period={freq};PriceAdj=F;TradingCalendar={calendar};showblank=0")
        # 仍然需要替换最后月末的价格
        end_next = w.tdaysoffset(1, end_, "TradingCalendar=%s" % calendar).Data[0][
            0].strftime("%Y-%m-%d")
        res1 = w.wsd(benchmark, "close", end_, end_next,
                    "PriceAdj=F;showblank=0;TradingCalendar=%s" % calendar)
        res.Data[0][-1] = res1.Data[0][0]
        rt_seq = close2return(res.Data, percentage=percentage, tolist=tolist, fillna_value=0.0)
        assert (len(res.Times) - 1) == len(time_seq), "TimeIndexError: 预测区间长度与实际获取长度不一致"
    else:
        end_ = time_seq[-1][:-4]+"1"+time_seq[-1][-3:]
        if pre_backtest_t:
            start_ = pre_backtest_t[:-4] + "1" + pre_backtest_t[-3:]
        else:
            start_ = time_seq[0]
        res = w.wsi(benchmark, "pct_chg", start_, end_,
                    f"BarSize=60;TradingCalendar={calendar};showblank=0.0")
        # DEBUG,可能存在重复的行，因此用pd.DataFrame结构去重，保留第一次出现的行
        remove_dup = pd.DataFrame(np.array(res.Data).T, index=res.Times, columns=[benchmark])
        remove_dup.reset_index(inplace=True)
        remove_dup.drop_duplicates("index", keep="first", inplace=True)
        remove_dup["index"] = [tt.strftime("%Y-%m-%d %H:%M:%S") for tt in remove_dup["index"].tolist()]
        # DEBUG，可能存在数据缺失，尝试重新获取修复，如果获取失败，用0.0 替代
        # 对比组合收益时间
        t_leak = list(set(time_seq) - set(remove_dup["index"].values.tolist()))
        if t_leak:
            t_leak = sorted(t_leak)
            t_leak[-1] = t_leak[-1][:-4]+"1"+t_leak[-1][-3:]
            leak_pad = w.wsi(benchmark, "pct_chg", t_leak[0], t_leak[-1],
                             f"BarSize=60;TradingCalendar={calendar};showblank=0.0")
            if leak_pad.ErrorCode != 0:
                leak_pad_df = pd.DataFrame({"index": t_leak, benchmark: [0.0]*len(t_leak)})
            else:
                leak_pad_df = pd.DataFrame({"index": t_leak, benchmark: leak_pad.Data[0]})
            remove_dup = pd.concat([remove_dup, leak_pad_df], ignore_index=True).sort_values(by=[
                "index"])
        res.Times = remove_dup["index"].tolist()
        res.Data = [remove_dup[benchmark].values.tolist()]
        assert len(res.Times) == len(time_seq) * rebalance_hour, \
            f"TimeIndexError:{benchmark}收益率序列可能存在缺失"
        rt_seq_time = res.Times
        rt_seq = pd.Series(res.Data[0], index=rt_seq_time)
        rt_seq = rt_seq / 100
        if rebalance_hour > 1:
            rt_seq = rt_seq.rolling(rebalance_hour).apply(lambda x: np.prod(x + 1) - 1)
            rt_seq = rt_seq.loc[time_seq]
        if percentage:
            rt_seq = rt_seq * 100
        if tolist:
            rt_seq = rt_seq.tolist()
    return rt_seq


def get_portfolio_basic_K_valuess(rts_arr, weights, time_seq, init_port_value, percentage=True,
                         tolist=False):
    """
    获取一段调仓周期内组合净值。

    :param rts_arr: 2Dlist, 各股票收益率
    :param weights: list, 最近调仓日权重
    :param time_seq: list, rts_arr对应时间
    :param init_port_value: float, 期初组合净值；基期为1
    :param percentage: bool, rts_arr是否是百分数收益率
    :param tolist: bool， 是否以list形式返回

    :return: pd.Series or list, float-->组合净值序列，期末净值
    """
    rts_df = pd.DataFrame(np.array(rts_arr).astype(np.float32), index=time_seq) / [1.0, 100][int(
        percentage)]
    cum_df = rts_df.apply(lambda x: (1+x).cumprod(), axis=0)
    cum_port_df = cum_df.apply(lambda x: np.dot(weights, x), axis=1)
    cum_port_df = cum_port_df * init_port_value
    init_port_value = cum_port_df.iloc[-1]
    return cum_port_df, init_port_value


if __name__ == "__main__":
    import os
    # import logging
    # logging.basicConfig(level=logging.DEBUG)
    try:
        wpc = WindPyChecker()
        if wpc.start_flag:
            wpc.wait_for_wsi(execute=True, wait_sec=5)
            m = Manage()
            m.run_optimizer()
    except Exception as e:
        print(e)
        traceback.print_exc()
    finally:
        os.system("pause")








