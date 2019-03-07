# -*- coding:utf-8 -*-
# ! python3

from configparser import ConfigParser
from Mkv_constant import *
from os import path, makedirs, mkdir
# import win32com.client as wc
from openpyxl import load_workbook, Workbook
from xlrd import open_workbook
from Mkv_data2 import create_stocks, calc_params
from Mkv_optimize2 import optimizer
from Mkv_pointwise_dates import BTdate
from Mkv_eval import MkvEval
from datetime import date, datetime
from WindPy import w


class Manage:
    def __init__(self):
        """
        设定运行基本信息
        """
        self.code_dir = ""
        self.work_dir = ""
        self.calc_t = {}
        self.mode = 2
        self.frequency = "M"
        self.start_t = ""
        self.end_t = ""
        # self.first_t = ""
        self.num_d = T
        self.up = DEFAULT_UP
        self.cash_r = CASH_RETURN
        self.cons_vol = DEFAULT_VOL / N ** 0.5
        self.calc_now = False
        self.short = None
        self.codes = []
        self.stocks = []
        self.stocks_panel = {}
        self.t_seq = []
        self.t_eval_seq = []
        self.w = []
        self.w_panel = {}
        self.params = None
        self.mu_panel = {}
        self.cov_panel = {}
        self.conf = ConfigParser()
        self.init_conf()
        self.set_code_file()
        self.set_work_file()
        self.set_mode()
        self.set_cons_vol()
        self.set_cons_up()
        self.set_cash_return()
        self.set_num_d()
        self.set_short()
        self.conf.write(open(CONF_NAME, "w+"))
        self.read_codes()

    def init_conf(self):
        """
        初始化参数配置文件
        :return:
        """
        f = open(CONF_NAME, "a+")
        f.close()
        self.conf.read(CONF_NAME, encoding="gbk")
        # self.conf.read_file(codecs.open(CONF_NAME, "r", encoding="gbk"))
        for sec in CONF_DICT.keys():
            if not self.conf.has_section(sec):
                self.conf.add_section(sec)
                for opt in CONF_DICT[sec]:
                    self.conf.set(sec, opt, "")
            else:
                for opt in CONF_DICT[sec]:
                    if not self.conf.has_option(sec, opt):
                        self.conf.set(sec, opt, "")
        print(f"-*-欢迎使用Markovitz组合优化系统 {VERSION}-*-")

    def set_code_file(self):
        """
        从键盘读取文件地址
        :return:
        """
        code_dir = self.conf.get("dir", "code_file")
        print("Part 1: 地址参数")
        print("1.1 股票代码文件地址")
        print(f"-当前股票代码文件地址：{code_dir}")
        code_dir_input = input("-请输入股票代码文件【不需修改请直接回车】：\n")
        flag = True
        while flag:
            if not code_dir_input:
                if not code_dir:
                    print(FILE_EMPTY_MSG)
                else:
                    self.code_dir = code_dir
                    flag = False
            elif not path.isfile(code_dir_input):
                code_dir_input = input(FILE_EXIST_MSG)
            else:
                self.code_dir = code_dir_input
                flag = False
                self.conf.set("dir", "code_file", self.code_dir)
        return

    def set_work_file(self):
        """
        从键盘读入工作文件夹地址
        :return:
        """
        print("1.2 工作文件夹地址")
        work_dir = self.conf.get("dir", "work_file")
        print(f"-当前工作文件夹地址：{work_dir}")
        work_dir_input = input(f"-请输入工作文件夹地址【不需修改请直接回车】：\n").strip("\r")
        if not work_dir_input:
            if "/Mkv_output" not in work_dir:
                work_dir += "/Mkv_output"
            if not path.exists(work_dir):
                makedirs(work_dir)
                self.work_dir = work_dir
            else:
                self.work_dir = work_dir
        elif not path.exists(work_dir_input):
            if "/Mkv_output" not in work_dir_input:
                work_dir_input += "/Mkv_output"
            makedirs(work_dir_input)
            print(f"该文件夹不存在，已为您新建{work_dir_input}")
            self.work_dir = work_dir_input
        else:
            if "/Mkv_output" not in work_dir_input:
                work_dir_input += "/Mkv_output"
            if not path.exists(work_dir_input):
                mkdir(work_dir_input)
                self.work_dir = work_dir_input
            else:
                self.work_dir = work_dir_input
        self.conf.set("dir", "work_file", self.work_dir)

    def set_mode(self):
        """
        设置运行模式
        :return:
        """
        print("Part 2: 运行参数")
        mode = self.conf.get("constraints", "mode")
        print(f"当前运行模式：mode={mode}")
        mode_input = input("请选择运行模式【1=回测模式，2=今日实盘模式，无需更改请直接回车】:\n").strip("\n\r")
        while True:
            if not mode_input or mode_input not in ["1", "2"]:
                if not mode or mode not in ["1", "2"]:
                    print("mode参数设置错误，请重新设置")
                    mode_input = input("请选择运行模式【1=回测模式，2=今日实盘模式，无需更改请直接回车】:\n").strip("\n\r")
                else:
                    self.mode = int(mode)
                    break
            else:
                self.mode = int(mode_input)
                self.conf.set("constraints", "mode", mode_input)
                break
        if self.mode == 1:
            self.set_time_interval()
        else:
            self.set_calc_time()

    def set_time_interval(self):
        print("2.1 回测时间区间")
        start_t = self.conf.get("calculation", "start_time")
        print(f"-当前回测开始时间设定：{start_t}")
        start_t_input = input("-请输入回测开始时间，格式yyyy-mm-dd【不需修改请直接回车】：\n").strip("\n\r")
        if not start_t_input:
            if not start_t:
                self.start_t = DEFAULT_START_T
            else:
                self.start_t = start_t
        else:
            self.start_t = start_t_input
            self.conf.set("calculation", "start_time", self.start_t)
        end_t = self.conf.get("calculation", "end_time")
        print(f"-当前回测结束时间设定：{end_t}")
        end_t_input = input("-请输入回测结束时间，格式yyyy-mm-dd【不需修改请直接回车】：\n").strip("\n\r")
        if not end_t_input:
            if not end_t:
                self.end_t = DEFAULT_END_T
            else:
                self.end_t = end_t
        else:
            self.end_t = end_t_input
            self.conf.set("calculation", "end_time", self.end_t)

        self.set_frequency()

    def set_frequency(self):
        """设置回测频率"""
        print("2.2.1 回测频率")
        freq = self.conf.get("calculation", "frequency")
        print(f"当前回测频率设定：{freq}")
        freq_input = input(
            "请输入回测时间间隔/频率【M-每月回测，W-每周回测，D-每日回测】：\n").strip("\n\r")
        if not freq_input:
            if freq:
                self.frequency = freq
        else:
            assert isinstance(freq_input, str) and freq_input.upper() in ["M", "W", "D"], \
                "输入回测频率'%s'不支持" % freq_input
            self.frequency = freq_input
            self.conf.set("calculation", "frequency", freq_input)


    def set_calc_time(self):
        """
        设置当日盘中价格截止时间
        :return:
        """

        print("2.1 价格截止时间")
        calc_time = self.conf.get("calculation", "calc_time")
        print(f"-当前截止时间设定：{calc_time}")
        calc_time_input = input(
            "-请输入计算截止时间，格式yyyy-mm-dd HH:MM:SS【不需修改请直接回车】：\n").strip("\n\r")
        if not calc_time_input:
            if not calc_time:
                self.calc_now = True
            else:
                from re import split
                t = [i.strip() for i in split("[:\- ]", calc_time)]
                # 补齐数组长度，防止用户漏写
                t = t + ["00"] * (6 - len(t))
                self.calc_t = {"y": int(t[0]), "m": int(t[1]), "d": int(t[2]),
                               "H": int(t[3]), "M": int(t[4]), "S": int(t[5])}
        else:
            from re import split
            t = [i.strip() for i in split("[:\- ]", calc_time_input)]
            t = t + ["00"] * (6 - len(t))
            self.calc_t = {"y": int(t[0]), "m": int(t[1]), "d": int(t[2]),
                           "H": int(t[3]), "M": int(t[4]), "S": int(t[5])}
            self.conf.set("calculation", "calc_time", calc_time_input)

    def set_cons_vol(self):
        """
        设置优化的年化波动率约束
        :return:
        """
        print("2.2 年化波动率约束")
        cons_vol = self.conf.get("constraints", "vol").strip()
        print(f"-当前年化波动率约束：{cons_vol}")
        cons_vol_input = input(
            "-请输入年化波动率约束【请用小数表示，如0.15;不需修改请直接回车】（<=）\n").strip("\r\n")
        if not cons_vol_input:
            if not cons_vol:
                self.cons_vol = DEFAULT_VOL / N ** 0.5
            else:
                self.cons_vol = float(cons_vol) / N ** 0.5
        else:
            self.cons_vol = float(cons_vol_input) / N ** 0.5
            self.conf.set("constraints", "vol", cons_vol_input)

    def set_cons_up(self):
        print("2.3 单股最大权重约束")
        up = self.conf.get("constraints", "max_weight").strip("")
        print(f'-当前单股最大权重约束：{up}')
        up_input = input("请输入单股最大权重（现金资产不受此约束，包括多头和空头）【不需修改请直接回车】：\n").strip("\r\n")
        if not up_input:
            if up:
                self.up = float(up)
        else:
            self.up = float(up_input)
            self.conf.set("constraints", "max_weight", up_input)

    def set_cash_return(self):
        print("2.4 现金年化收益率")
        cash_r = self.conf.get("constraints", "cash_return").strip("")
        print(f"-当前设置的现金年化收益率：{cash_r}")
        cash_r_input = input("请输入现金年化收益率【不需修改请直接回车】：\n").strip("\r\n")
        if not cash_r_input:
            if cash_r:
                self.cash_r = float(cash_r) / 360
        else:
            self.cash_r = float(cash_r_input) / 360
            self.conf.set("constraints", "cash_return", cash_r_input)

    def set_num_d(self):
        print("2.4 历史收益率交易日长度")
        num_d = self.conf.get("calculation", "num_d")
        print(f"-当前收益率历史交易日长度：{num_d}")
        num_d_input = input(
            "-请输入历史收益率交易日长度【不需修改请直接回车】：\n"
        )
        if not num_d_input:
            if num_d:
                self.num_d = int(num_d)
        else:
            self.num_d = int(num_d_input)
            self.conf.set("calculation", "num_d", str(self.num_d))

    def set_short(self):
        print("2.5 做空约束")
        short = self.conf.get("constraints", "short")
        print(f"-当前做空约束：{short}")
        short_input = input(
            "-请选择是否允许空头头寸，1=允许做空，0=不允许做空【不需修改请直接回车】：\n"
        ).strip()
        if not short_input:
            if not short:
                self.short = 0
                print(SHORT_MISS_WARNING)
            else:
                self.short = int(short)
        else:
            self.short = int(short_input)
            self.conf.set("constraints", "short", short_input)

    def read_codes(self):
        """
        从文件中读取代码
        :return:
        """
        print("*************************************************************")
        self.codes = read_codes(self.code_dir)
        print(f"%读取{len(self.codes)}只股票")

    def set_params1(self):
        """
        回测模式的优化计算
        :return:
        """
        for tt in self.t_seq:
            print(f"-正在回测时点{tt}")
            stocks = create_stocks(self.codes, self.num_d, tt, False)
            self.stocks_panel.update({tt: stocks})
            params = calc_params(stocks, self.short)
            params.update({"vol": self.cons_vol, "up": self.up, "cash_r": self.cash_r})
            self.mu_panel.update({tt: params["mu"]})
            self.cov_panel.update({tt: params["cov"]})
            w = optimizer(params)
            self.w_panel.update({tt: w})
            import numpy as np
            print(f"--Backtest[{tt}] output:{np.matmul(np.matmul(np.array(w), params['cov']), np.array(w).T)}")
            print(f"--constraint:{self.cons_vol**2}")
        self.write_output1()

    def set_params2(self, q=True):
        """
        实盘问题的计算优化参数
        :return:
        """
        if not q:
            self.stocks = create_stocks(self.codes,
                                        self.num_d,
                                        datetime(*self.calc_t.values()).strftime("%Y-%m-%d"),
                                        q=q)
        else:
            self.stocks = create_stocks(self.codes, self.num_d)
        self.params = calc_params(self.stocks, self.short)
        self.params.update({"vol": self.cons_vol, "up": self.up, "cash_r": self.cash_r})
        self.w = optimizer(self.params)
        import numpy as np
        print(
            f"output:"
            f"{np.matmul(np.matmul(np.array(self.w), self.params['cov']), np.array(self.w).T)}\n"
            f"constraint:{self.cons_vol**2}")
        self.write_output2()

    def get_backtest_t_seq(self):
        bt_obj = BTdate(start=self.start_t,
                        end=self.end_t,
                        freq=self.frequency)
        self.t_seq = bt_obj.get_backtest_dates()
        self.t_eval_seq = bt_obj.get_eval_dates(ref_seq=self.t_seq)
        # self.first_t = bt_obj.get_backtest_first_date(num_d=self.num_d)

    def run_optimizer(self):
        if self.mode == 1:
            self.get_backtest_t_seq()
            print(f"-*-进入回测模式，检测到{len(self.t_seq)}个回测时点-*-")
            self.set_params1()
        else:
            stamp = datetime.now()
            calc_t = datetime(*self.calc_t.values())
            # 给定日期非当天，用于回测而非实盘
            if calc_t.date() < stamp.date():
                self.set_params2(q=False)
            # 给定日期在当天或者在未来，都当做在当天进行
            elif calc_t.date() > stamp.date():
                self.set_params2()
            else:
                if calc_t <= stamp:
                    self.set_params2()
                else:
                    delta_t = (calc_t - stamp).seconds
                    print("%定时等待运行...")
                    from time import sleep
                    sleep(delta_t)
                    self.set_params2()

    def write_output1(self):
        wb = Workbook()
        s1 = wb.active
        s1.title = "weights"
        s2 = wb.create_sheet()
        token = {"M": "monthly", "W": "weekly", "D": "daily"}[self.frequency]
        s2.title = "portfolio %s returns" % token
        s3 = wb.create_sheet()
        s3.title = f"{self.num_d}-day mean returns"
        # 写第一张表回测各周期weights
        col11 = ["code"] + self.codes + ["cash"]
        names = [s.name for s in self.stocks_panel[self.t_seq[0]]]
        col12 = ["name"] + names + ["现金"]
        content1 = [col11, col12]
        for tt in self.t_seq:
            content1.append([tt] + self.w_panel[tt])
        content1 = list(zip(*content1))
        for row in content1:
            s1.append(row)
        # 写第二张表回测权重持有至下一期的组合期间收益率
        eval_obj = MkvEval(eval_seq=self.t_eval_seq,
                           freq=self.frequency,
                           stocks=self.codes,
                           cash_r=self.cash_r)
        returns = eval_obj.eval_returns(var="pct_chg")
        from numpy import dot
        col22 = ["portfolio return(%)"] + \
                [dot(self.w_panel[self.t_seq[t]], returns[t]) for t in
                                         range(len(self.t_seq))]
        s2.append(["evaluation_date"] + self.t_eval_seq)
        s2.append(col22)
        # 写第三张表
        content3 = [col11, col12]
        for tt in self.t_seq:
            content3.append([tt] + self.mu_panel[tt] + [self.cash_r])
        content3 = list(zip(*content3))
        for row in content3:
            s3.append(row)

        # 写回测期间协方差的表
        for t in range(len(self.t_seq)):
            exec(f"s{4+t}=wb.create_sheet()")
            exec(f"s{4+t}.title='Cov_{self.t_seq[t]}'")
            for row in self.cov_panel[self.t_seq[t]]:
                exec(f"s{4+t}.append({list(row)})")

        output_dir = f"{self.work_dir}/mkv_{path.splitext(path.basename(self.code_dir))[0]}"+ \
            f"_{['longOnly', 'short'][self.short]}_" \
            f"{''.join(self.start_t.split('-'))}-{''.join(self.end_t.split('-'))}.xlsx"
        wb.save(output_dir)
        print(f"%结果输出到：{output_dir}")

    def write_output2(self):
        wb = Workbook()
        s1 = wb.active
        s1.title = "weights"
        s2 = wb.create_sheet()
        s2.title = f"{self.num_d}-days mean returns"
        s3 = wb.create_sheet()
        s3.title = "covariance"
        # 写第一张表-weight
        s1.cell(row=1, column=1, value="code")
        s1.cell(row=1, column=2, value="name")
        s1.cell(row=1, column=3, value="weight")
        for cc in range(len(self.stocks)):
            s1.cell(row=cc + 2, column=1, value=self.stocks[cc].code)
            s1.cell(row=cc + 2, column=2, value=self.stocks[cc].name)
            s1.cell(row=cc + 2, column=3, value=self.w[cc])
        s1.cell(row=len(self.w) + 1, column=1, value="cash")
        s1.cell(row=len(self.w) + 1, column=2, value="现金")
        s1.cell(row=len(self.w) + 1, column=3, value=self.w[-1])
        # 写第二张表-return
        s2.cell(row=1, column=1, value="code")
        s2.cell(row=1, column=2, value="name")
        s2.cell(row=1, column=3, value="mu")
        for cc in range(len(self.stocks)):
            s2.cell(row=cc + 2, column=1, value=self.stocks[cc].code)
            s2.cell(row=cc + 2, column=2, value=self.stocks[cc].name)
            s2.cell(row=cc + 2, column=3, value=self.params["mu"][cc])
            for t in range(len(self.stocks[0].t)):
                s2.cell(row=1, column=t + 4, value=self.stocks[0].t[t])
                s2.cell(row=cc + 2, column=t + 4, value=self.stocks[cc].r[t])
        s2.cell(row=len(self.w) + 1, column=1, value="cash")
        s2.cell(row=len(self.w) + 1, column=2, value="现金")
        s2.cell(row=len(self.w) + 1, column=3, value=0.)
        for t in range(len(self.stocks[0].t)):
            s2.cell(row=len(self.w) + 1, column=t + 4, value=self.cash_r)
        # 写第三张表-covariance
        s3.cell(row=1, column=1, value="code")
        s3.cell(row=1, column=2, value="name")
        for cc in range(len(self.stocks)):
            s3.cell(row=cc + 2, column=1, value=self.stocks[cc].code)
            s3.cell(row=cc + 2, column=2, value=self.stocks[cc].name)
        for i in range(len(self.params['qval'])):
            s3.cell(row=self.params['qsubi'][i] + 2,
                    column=self.params['qsubj'][i] +3,
                    value=self.params['qval'][i])
        output_dir = f"{self.work_dir}/mkv_{path.splitext(path.basename(self.code_dir))[0]}"+ \
            f"_{['longOnly', 'short'][self.short]}_{datetime(*self.calc_t.values()).strftime('%y%m%d')}.xlsx"
        wb.save(output_dir)
        print(f"%结果输出到：{output_dir}")


def read_codes(f):
    b_name = path.basename(f)
    b_type = b_name.split(".")[-1]
    codes = []
    # if b_type == 'xls':
    #     # openpyxl不支持xls格式，先转化为xls
    #     e = wc.gencache.EnsureDispatch('Excel.Application')
    #     wb = e.Workbooks.Open(f)
    #     wb.SaveAs(f + 'x', FileFormat=51)
    #     wb.Close()
    #     e.Application.Quit()
    #     f = f + 'x'
    #     b_type = 'xlsx'
    if b_type == "txt":
        for code in open(f):
            code = code.strip("\n|;|,|/|")
            codes.append(code)
    elif b_type == "xlsx":
        wb = load_workbook(f)
        sheet = wb.active
        for code in list(sheet.columns)[0]:
            if code.value and "." in code.value:
                codes.append(code.value)
    elif b_type == 'xls':
        wb = open_workbook(f)
        sheet = wb.sheet_by_index(0)
        for code in sheet.col_values(0):
            if "." in code:
                codes.append(code)
    else:
        print(f"Error: 不支持的文件格式{b_type}")
    return codes


def test_read_codes(f="C:/Users/Maggie/Desktop/test.xls"):
    """
    read_codes函数的单元测试
    :param f: 文件地址
    :return:
    """
    res = read_codes(f)
    print(res)


if __name__ == "__main__":
    m = Manage()
    m.run_optimizer()








