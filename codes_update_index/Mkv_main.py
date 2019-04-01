# -*- coding:utf-8 -*-
# ! python3

from configparser import ConfigParser
from Mkv_constant import *
from os import path, makedirs, mkdir
# import win32com.client as wc
from openpyxl import load_workbook, Workbook
from xlrd import open_workbook
from Mkv_data2 import create_stocks, calc_params
from Mkv_optimize2 import optimizer
from Mkv_pointwise_dates import BTdate
from Mkv_pointwise_codes import BTcodes
from Mkv_eval import MkvEval
from datetime import date, datetime
import traceback
import pandas as pd
from WindPy import w


class Manage:
    def __init__(self):
        """
        设定运行基本信息
        """
        self.code_dir = ""
        self.work_dir = ""
        self.target_index = ""
        self.global_spec = ""
        self.calc_t = {}
        self.mode = 2
        self.frequency = "M"
        self.start_t = ""
        self.end_t = ""
        # self.first_t = ""
        self.num_d = T
        self.basic_indices = ""
        self.refresh_freq = ""
        self.up = DEFAULT_UP
        self.cash_r = CASH_RETURN
        self.cons_vol = DEFAULT_VOL / N ** 0.5
        self.calc_now = False
        self.short = None
        self.codes = []
        self.stocks = []
        self.stocks_panel = {}
        self.t_seq = []
        self.t_eval_seq = []
        self.w = []
        self.w_panel = {}
        self.params = None
        self.mu_panel = {}
        self.cov_panel = {}
        self.bt_obj = None
        self.conf = ConfigParser()
        self.init_conf()
        self.read_target_index()
        self.set_input_mode()
        self.read_work_file()
        self.read_mode()
        self.read_cons_vol()
        self.read_cons_up()
        self.read_cash_return()
        self.read_num_d()
        self.read_short()
        self.conf.write(open(CONF_NAME, "w+"))

    # ------------------------
    # 检验参数配置文件有效性
    # ------------------------
    def init_conf(self):
        """
        初始化参数配置文件
        :return:
        """
        f = open(CONF_NAME, "a+")
        f.close()
        self.conf.read(CONF_NAME, encoding="gbk")
        # self.conf.read_file(codecs.open(CONF_NAME, "r", encoding="gbk"))
        for sec in CONF_DICT.keys():
            if not self.conf.has_section(sec):
                self.conf.add_section(sec)
                for opt in CONF_DICT[sec]:
                    self.conf.set(sec, opt, "")
            else:
                for opt in CONF_DICT[sec]:
                    if not self.conf.has_option(sec, opt):
                        self.conf.set(sec, opt, "")
        print(f"-*-欢迎使用Markovitz组合优化系统 {VERSION}-*-")
        print("正在读取参数文件...")

    # ----------------------------------
    # 从configuration文件中读取参数设定
    # ----------------------------------
    def read_target_index(self):
        """
        从configuration文件中读取跟踪的目标指数目标指数
        :return:
        """
        self.target_index = self.conf.get("dir", "target_index")
        if not self.target_index:
            self.read_global_spec()

    def read_global_spec(self):
        """
        从configuration文件中读取基本面过滤全局股票空间，通常是某个全市场指数
        :return:
        """
        self.global_spec = self.conf.get("filter", "global_spec")
        if not self.global_spec:
            self.read_code_file()

    def read_code_file(self):
        """
        从configuration中读取参数设定
        :return:
        """
        self.code_dir = self.conf.get("dir", "code_file")
        flag = True
        while flag:
            if not path.isfile(self.code_dir):
                self.code_dir = input(FILE_EXIST_MSG)
            else:
                flag = False
                self.conf.set("dir", "code_file", self.code_dir)
        return

    def set_input_mode(self):
        if self.target_index:
            self.input_mode = 1
        elif self.global_spec:
            self.input_mode = 2
            self.read_indices()
            self.read_refresh_freq()
            self.read_ics()
            self.read_ics_fv()
            self.read_ics_rank()
        elif self.code_dir:
            self.input_mode = 3
        else:
            self.input_mode = None
            raise Exception("InputModeError: Can't decide stock pools.")

    def read_work_file(self):
        """
        从configuration读入工作文件夹地址
        :return:
        """
        self.work_dir = self.conf.get("dir", "work_file")

        if not path.exists(self.work_dir):
            if "/Mkv_output" not in self.work_dir:
                self.work_dir += "/Mkv_output"
            makedirs(self.work_dir)
            print(f"该文件夹不存在，已为您新建{self.work_dir}")

        else:
            if "/Mkv_output" not in self.work_dir:
                self.work_dir += "/Mkv_output"
                if not path.exists(self.work_dir):
                    mkdir(self.work_dir)
                    self.conf.set("dir", "work_file", self.work_dir)

    def read_mode(self):
        """
        读取并设置运行模式
        :return:
        """
        self.mode = self.conf.get("constraints", "mode")
        assert self.mode.isdigit(), "BacktestModeError:模式值不能为空，期望1=回测模式，2=实盘模式"
        self.mode = int(self.mode)
        assert self.mode in [1, 2], "BacktestModeError:模式值只能是1或2"
        if self.mode == 1:
            self.read_time_interval()
        else:
            self.read_calc_time()

    def read_time_interval(self):
        """
        从configuration文件中读入回测区间（起始时间）
        :return:
        """
        self.start_t = self.conf.get("calculation", "start_time")
        if not self.start_t:
            self.start_t = DEFAULT_START_T
            print(f"Warning:未输入回测起始时间，采用默认时间{DEFAULT_START_T}")
            self.conf.set("calculation", "start_time", self.start_t)

        self.end_t = self.conf.get("calculation", "end_time")
        if not self.end_t:
            self.end_t = DEFAULT_END_T
            print(f"Warning:未输入回测截止时间，采用默认时间{DEFAULT_END_T}")
            self.conf.set("calculation", "end_time", self.end_t)

        self.read_frequency()

    def read_frequency(self):
        """读取回测频率"""

        self.frequency = self.conf.get("calculation", "frequency")
        if self.mode != 2:
            assert isinstance(self.frequency, str) and self.frequency.upper() in ["M", "W", "D"], \
                "输入回测频率'%s'不支持" % self.frequency

    def read_calc_time(self):
        """
        设置当日盘中价格截止时间
        :return:
        """
        self.calc_time = self.conf.get("calculation", "calc_time")
        if not self.calc_time:
                self.calc_now = True
        else:
            from re import split
            t = [i.strip() for i in split("[:\- ]", self.calc_time)]
            # 补齐数组长度，防止用户漏写
            t = t + ["00"] * (6 - len(t))
            self.calc_t = {"y": int(t[0]), "m": int(t[1]), "d": int(t[2]),
                           "H": int(t[3]), "M": int(t[4]), "S": int(t[5])}

    def read_cons_vol(self):
        """
        读取优化的年化波动率约束
        :return:
        """
        cons_vol = self.conf.get("constraints", "vol").strip()
        if not cons_vol:
            self.cons_vol = DEFAULT_VOL / N ** 0.5
        else:
            self.cons_vol = float(cons_vol) / N ** 0.5

    def read_cons_up(self):
        """
        读取最大仓位约束
        :return:
        """
        self.up = float(self.conf.get("constraints", "max_weight").strip(""))

    def read_cash_return(self):
        """
        读取年化现金收益率
        :return:
        """
        self.cash_r = float(self.conf.get("constraints", "cash_return").strip("")) / 360

    def read_num_d(self):
        """
        读取回测日期长度
        :return:
        """
        self.num_d = int(self.conf.get("calculation", "num_d"))

    def read_short(self):
        """
        读取回测读取做空限制
        :return:
        """
        self.short = int(self.conf.get("constraints", "short"))

    def read_indices(self):
        """
        读取基本面指标字符串
        :return:
        """
        self.indices = self.conf.get("filter", "basic_indices")

    def read_refresh_freq(self):
        """
        读取更新筛选股票池更新频率
        :return:
        """
        self.refresh_freq = self.conf.get("filter", "refresh_freq").strip("\n\r").upper()
        if self.refresh_freq:
            assert int(self.refresh_freq[0]) in [1, 2, 3, 4, 5, 6] and self.refresh_freq[1] == \
                   "M", \
                "Error: parameter refresh_freq is not supported, should be in 1M~6M"

    def read_ics(self):
        """
        读取二级行业分类标准
        :return:
        """
        self.ics = self.conf.get("filter", "ics").strip("\n\r").lower()
        if not self.ics:
            self.ics = DEFAULT_ICS
            print("Warning: 参数ics（行业分类标准）未正确输入，采用默认输入--industry_gics（WIND行业分类）")

    def read_ics_fv(self):
        """
        读取子行业内排名变量，暂时只支持单个变量排序
        :return:
        """
        self.ics_fv = self.conf.get("filter", "ics_fv").strip("\n\r").lower()
        if not self.ics_fv:
            self.ics_fv = DEFAULT_ICS_FV
        else:
            self.ics_fv = FIELDS_ALIAS_DICT.get(self.ics_fv, 0)
            if not self.ics_fv:
                print("Warning：ics_fv（二级行业内排序变量）未正确输入，采用默认输入--ev（市值）")
                self.ics_fv = DEFAULT_ICS_FV

    def read_ics_rank(self):
        """
        读取子行业内排名变量，暂时只支持单个变量排序
        :return:
        """
        self.ics_rank = self.conf.get("filter", "ics_rank").strip("\n\r").lower()
        if not self.ics_rank:
            self.ics_rank = ICS_RANK
            print(f"Warning: 参数ics_rank（单个行业入选股票数量）未正确输入，采用默认输入rank={ICS_RANK}")
        elif not self.ics_rank.isdigit():
            raise Exception("ParameterError: 参数ics_rank必须为正整数")
        else:
            self.ics_rank = int(self.ics_rank)

    def read_codes(self):
        """
        从文件中读取代码
        :return:
        """
        print("*************************************************************")
        self.codes = read_codes(self.code_dir)
        print(f"%读取{len(self.codes)}只股票")

    def set_params1(self):
        """
        回测模式的优化计算
        :return:
        """
        if self.input_mode == 3:
            input_params = {"code_file": self.code_dir}
        elif self.input_mode == 1:
            input_params = {"target_index": self.target_index}
        else:
            input_params = {"global_spec": self.global_spec,
                            "basic_indices": self.indices,
                            "ics": self.ics,
                            "ics_fv": self.ics_fv,
                            "ics_rank": self.ics_rank,
                            "refresh_freq": self.refresh_freq}
        codes_object = BTcodes(input_mode=self.input_mode)
        codes_object.set_codes_env(params=input_params)
        for tt, month_id in self.t_seq:
            print(f"-正在回测时点{tt}")
            codes = codes_object.get_current_codes(date=tt, month_id=month_id)
            stocks = create_stocks(codes, self.num_d, tt, False)
            self.stocks_panel.update({tt: stocks})
            params = calc_params(stocks, self.short)
            params.update({"vol": self.cons_vol, "up": self.up, "cash_r": self.cash_r})
            self.mu_panel.update({tt: params["mu"]})
            self.cov_panel.update({tt: params["cov"]})
            w = optimizer(params)
            self.w_panel.update({tt: w})
            import numpy as np
            print(f"--Backtest[{tt}] output:{np.matmul(np.matmul(np.array(w), params['cov']), np.array(w).T)}")
            print(f"--constraint:{self.cons_vol**2}")
        self.t_seq = [tt[0] for tt in self.t_seq]
        self.t_eval_seq = self.bt_obj.get_eval_dates(ref_seq=self.t_seq)
        if self.input_mode == 3:
            self.write_output1()
        else:
            self.write_output1d()

    def set_params2(self, q=True):
        """
        实盘问题的计算优化参数
        :return:
        """
        if self.input_mode == 3:
            input_params = {"code_file": self.code_dir}
        elif self.input_mode == 1:
            input_params = {"target_index": self.target_index}
        else:
            input_params = {"global_spec": self.global_spec,
                            "basic_indices": self.indices,
                            "ics": self.ics,
                            "ics_fv": self.ics_fv,
                            "ics_rank": self.ics_rank,
                            "refresh_freq": self.refresh_freq}
        codes_object = BTcodes(input_mode=self.input_mode)
        codes_object.set_codes_env(params=input_params)
        self.codes = codes_object.get_current_codes(date=datetime.now().strftime("%Y-%m-%d"),
                                               month_id=0)
        if not q:
            self.stocks = create_stocks(self.codes,
                                        self.num_d,
                                        datetime(*self.calc_t.values()).strftime("%Y-%m-%d"),
                                        q=q)
        else:
            self.stocks = create_stocks(self.codes, self.num_d)
        self.params = calc_params(self.stocks, self.short)
        self.params.update({"vol": self.cons_vol, "up": self.up, "cash_r": self.cash_r})
        self.w = optimizer(self.params)
        import numpy as np
        print(
            f"output:"
            f"{np.matmul(np.matmul(np.array(self.w), self.params['cov']), np.array(self.w).T)}\n"
            f"constraint:{self.cons_vol**2}")
        self.write_output2()

    def get_backtest_t_seq(self):
        self.bt_obj = BTdate(start=self.start_t,
                        end=self.end_t,
                        freq=self.frequency)
        self.t_seq = self.bt_obj.get_backtest_dates()
        # self.first_t = self.bt_obj.get_backtest_first_date(num_d=self.num_d)

    def run_optimizer(self):
        """
        根据运行模式的不同，选择回测或者是实盘函数入口
        :return:
        """
        if self.mode == 1:
            self.get_backtest_t_seq()
            print(f"-*-进入回测模式，检测到{len(self.t_seq)}个回测时点-*-")
            self.set_params1()
        else:
            stamp = datetime.now()
            calc_t = datetime(*self.calc_t.values())
            # 给定日期非当天，用于回测而非实盘
            if calc_t.date() < stamp.date():
                self.set_params2(q=False)
            # 给定日期在当天或者在未来，都当做在当天进行
            elif calc_t.date() > stamp.date():
                self.set_params2()
            else:
                if calc_t <= stamp:
                    self.set_params2()
                else:
                    delta_t = (calc_t - stamp).seconds
                    print("%定时等待运行...")
                    from time import sleep
                    sleep(delta_t)
                    self.set_params2()

    def write_output1(self):
        print("%正在写入结果到文件...")
        wb = Workbook()
        s1 = wb.active
        s1.title = "weights"
        s2 = wb.create_sheet()
        token = {"M": "monthly", "W": "weekly", "D": "daily"}[self.frequency]
        s2.title = "portfolio %s returns" % token
        s3 = wb.create_sheet()
        s3.title = f"{self.num_d}-day mean returns"
        # 写第一张表回测各周期weights
        col11 = ["code"] + self.codes + ["cash"]
        names = [s.name for s in self.stocks_panel[self.t_seq[0]]]
        col12 = ["name"] + names + ["现金"]
        content1 = [col11, col12]
        for tt in self.t_seq:
            content1.append([tt] + self.w_panel[tt])
        content1 = list(zip(*content1))
        for row in content1:
            s1.append(row)
        # 写第二张表回测权重持有至下一期的组合期间收益率
        eval_obj = MkvEval(eval_seq=self.t_eval_seq,
                           freq=self.frequency,
                           stocks=self.codes,
                           cash_r=self.cash_r)
        returns = eval_obj.eval_returns(var="pct_chg")
        from numpy import dot
        col22 = ["portfolio return(%)"] + \
                [dot(self.w_panel[self.t_seq[t]], returns[t]) for t in
                                         range(len(self.t_seq))]
        s2.append(["evaluation_date"] + self.t_eval_seq)
        s2.append(col22)
        # 写第三张表
        content3 = [col11, col12]
        for tt in self.t_seq:
            content3.append([tt] + self.mu_panel[tt] + [self.cash_r])
        content3 = list(zip(*content3))
        for row in content3:
            s3.append(row)

        # 写回测期间协方差的表
        for t in range(len(self.t_seq)):
            exec(f"s{4+t}=wb.create_sheet()")
            exec(f"s{4+t}.title='Cov_{self.t_seq[t]}'")
            for row in self.cov_panel[self.t_seq[t]]:
                exec(f"s{4+t}.append({list(row)})")

        output_dir = f"{self.work_dir}/mkv_{path.splitext(path.basename(self.code_dir))[0]}"+ \
            f"_{['longOnly', 'short'][self.short]}_" \
            f"{''.join(self.start_t.split('-'))}-{''.join(self.end_t.split('-'))}_freq{self.frequency}.xlsx"
        wb.save(output_dir)
        print(f"%结果输出到：{output_dir}")

    def write_output1d(self):
        print("%正在写入结果到文件...")
        wb = Workbook()
        s1 = wb.active
        s1.title = "weights"
        s2 = wb.create_sheet()
        token = {"M": "monthly", "W": "weekly", "D": "daily"}[self.frequency]
        s2.title = "portfolio %s returns" % token
        s3 = wb.create_sheet()
        s3.title = f"{self.num_d}-day mean returns"
        # 写第一张表回测各周期weights
        content1 = dict2list(stock_pannel=self.stocks_panel, var_pannel=self.w_panel)
        for row in content1:
            s1.append(row)
        # 写第二张表回测权重持有至下一期的组合期间收益率
        eval_obj = MkvEval(eval_seq=self.t_eval_seq,
                           freq=self.frequency,
                           stocks=[],
                           cash_r=self.cash_r)
        from numpy import dot
        col22 = ["portfolio return(%)"]
        for t in range(len(self.t_eval_seq)):
            codes = [s.code for s in self.stocks_panel[self.t_seq[t]]]
            returns = eval_obj.eval_returns12(codes=codes,
                                              eval_date=self.t_eval_seq[t],
                                              var="pct_chg")

            col22.append(dot(self.w_panel[self.t_seq[t]], returns))
        s2.append(["evaluation_date"] + self.t_eval_seq)
        s2.append(col22)
        # 写第三张表
        content3 = dict2list(stock_pannel=self.stocks_panel,
                             var_pannel=self.mu_panel,
                             suffix=[self.cash_r])
        for row in content3:
            s3.append(row)
        # 写回测期间协方差的表
        for t in range(len(self.t_seq)):
            exec(f"s{4 + t}=wb.create_sheet()")
            exec(f"s{4 + t}.title='Cov_{self.t_seq[t]}'")
            for row in self.cov_panel[self.t_seq[t]]:
                exec(f"s{4 + t}.append({list(row)})")

        output_dir = f"{self.work_dir}/mkv_{[self.target_index, self.global_spec][self.input_mode-1]}" + \
                     f"_{['longOnly', 'short'][self.short]}_" \
                         f"{''.join(self.start_t.split('-'))}-{''.join(self.end_t.split('-'))}_freq{self.frequency}.xlsx"
        wb.save(output_dir)
        print(f"%结果输出到：{output_dir}")

    def write_output2(self):
        print("%正在写入结果到文件...")
        wb = Workbook()
        s1 = wb.active
        s1.title = "weights"
        s2 = wb.create_sheet()
        s2.title = f"{self.num_d}-days mean returns"
        s3 = wb.create_sheet()
        s3.title = "covariance"
        # 写第一张表-weight
        s1.cell(row=1, column=1, value="code")
        s1.cell(row=1, column=2, value="name")
        s1.cell(row=1, column=3, value="weight")
        for cc in range(len(self.stocks)):
            s1.cell(row=cc + 2, column=1, value=self.stocks[cc].code)
            s1.cell(row=cc + 2, column=2, value=self.stocks[cc].name)
            s1.cell(row=cc + 2, column=3, value=self.w[cc])
        s1.cell(row=len(self.w) + 1, column=1, value="cash")
        s1.cell(row=len(self.w) + 1, column=2, value="现金")
        s1.cell(row=len(self.w) + 1, column=3, value=self.w[-1])
        # 写第二张表-return
        s2.cell(row=1, column=1, value="code")
        s2.cell(row=1, column=2, value="name")
        s2.cell(row=1, column=3, value="mu")
        for cc in range(len(self.stocks)):
            s2.cell(row=cc + 2, column=1, value=self.stocks[cc].code)
            s2.cell(row=cc + 2, column=2, value=self.stocks[cc].name)
            s2.cell(row=cc + 2, column=3, value=self.params["mu"][cc])
            for t in range(len(self.stocks[0].t)):
                s2.cell(row=1, column=t + 4, value=self.stocks[0].t[t])
                s2.cell(row=cc + 2, column=t + 4, value=self.stocks[cc].r[t])
        s2.cell(row=len(self.w) + 1, column=1, value="cash")
        s2.cell(row=len(self.w) + 1, column=2, value="现金")
        s2.cell(row=len(self.w) + 1, column=3, value=0.)
        for t in range(len(self.stocks[0].t)):
            s2.cell(row=len(self.w) + 1, column=t + 4, value=self.cash_r)
        # 写第三张表-covariance
        s3.cell(row=1, column=1, value="code")
        s3.cell(row=1, column=2, value="name")
        for cc in range(len(self.stocks)):
            s3.cell(row=cc + 2, column=1, value=self.stocks[cc].code)
            s3.cell(row=cc + 2, column=2, value=self.stocks[cc].name)
        for i in range(len(self.params['qval'])):
            s3.cell(row=self.params['qsubi'][i] + 2,
                    column=self.params['qsubj'][i] +3,
                    value=self.params['qval'][i])
        output_dir = f"{self.work_dir}/mkv_{[self.target_index, self.global_spec][self.input_mode-1]}"+ \
            f"_{['longOnly', 'short'][self.short]}_{datetime(*self.calc_t.values()).strftime('%y%m%d')}.xlsx"
        wb.save(output_dir)
        print(f"%结果输出到：{output_dir}")


def read_codes(f):
    b_name = path.basename(f)
    b_type = b_name.split(".")[-1]
    codes = []
    # if b_type == 'xls':
    #     # openpyxl不支持xls格式，先转化为xls
    #     e = wc.gencache.EnsureDispatch('Excel.Application')
    #     wb = e.Workbooks.Open(f)
    #     wb.SaveAs(f + 'x', FileFormat=51)
    #     wb.Close()
    #     e.Application.Quit()
    #     f = f + 'x'
    #     b_type = 'xlsx'
    if b_type == "txt":
        for code in open(f):
            code = code.strip("\n|;|,|/|")
            codes.append(code)
    elif b_type == "xlsx":
        wb = load_workbook(f)
        sheet = wb.active
        for code in list(sheet.columns)[0]:
            if code.value and "." in code.value:
                codes.append(code.value)
    elif b_type == 'xls':
        wb = open_workbook(f)
        sheet = wb.sheet_by_index(0)
        for code in sheet.col_values(0):
            if "." in code:
                codes.append(code)
    else:
        print(f"Error: 不支持的文件格式{b_type}")
    return codes


def test_read_codes(f="C:/Users/Maggie/Desktop/test.xls"):
    """
    read_codes函数的单元测试
    :param f: 文件地址
    :return:
    """
    res = read_codes(f)
    print(res)


# helper func: merge weights in different sets of stock pools
def dict2list(stock_pannel, var_pannel, suffix=None):
    """
    这个函数扩展性不佳，为了分隔较复杂的数据形式重组功能
    :param stock_pannel: 包含回测全部权股票基本信息的字典数据
    :param var_pannel：包含回测全部权重信息的字典数据
    :param suffix：需要附加在var_pannel元素后的list
    :return: 通过pd.DataFrame merge各个不同日期回测结果后转换为可以直接写入workbook的list形式
    """
    # 提取时间索引
    t_seq = list(stock_pannel.keys())
    t_start = t_seq[0]

    # 定义函数将每次回测结果转化为DataFrame
    def dict2dataframe(key):
        content0 = [[s.code, s.name] for s in stock_pannel[key]]
        content0.append(["cash", "现金"])
        content0 = dict(zip(["code", "name"],zip(*content0)))
        if suffix:
            if isinstance(var_pannel[key], list):
                var_pannel_ = var_pannel[key] + suffix
            else:
                var_pannel_ = list(var_pannel) + suffix
            content0.update({key: var_pannel_})
            df0 = pd.DataFrame(content0)
        else:
            content0.update({key: var_pannel[key]})
            df0 = pd.DataFrame(content0)
        return df0

    # 初始化总表
    df = dict2dataframe(t_start)
    # 通过循环，每次加入一次回测结果，保存之前结果merge后的数据框
    for tt in t_seq[1:]:
        df0 = dict2dataframe(tt)
        df = pd.merge(df, df0, on=["code", "name"], how="outer", copy=False)
    # 由于没有采用index，需要在输出之前调整行列顺序使其美观
    # 调整列顺序
    df = df[["code", "name"] + t_seq]
    # 调整行序，使得现金始终排在最后一行
    cash_id = df[df.code == "cash"].index.tolist()
    cash_row = df.iloc[cash_id]
    df.drop(index=cash_id, inplace=True)
    df = df.append(cash_row)
    df2list = df.values.tolist()
    df2list.insert(0, df.columns.tolist())
    return df2list


if __name__ == "__main__":
    import os
    # import logging
    # logging.basicConfig(level=logging.DEBUG)
    try:
        m = Manage()
        m.run_optimizer()
    except Exception as e:
        print(e)
        traceback.print_exc()
    finally:
        os.system("pause")








